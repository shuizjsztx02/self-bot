from typing import List, Dict, Any, Optional
from pathlib import Path
import asyncio

from .base import DocumentParser, ParsedDocument, ChunkResult


class ExcelParser(DocumentParser):
    
    async def parse(self, file_path: str) -> ParsedDocument:
        loop = asyncio.get_event_loop()
        return await loop.run_in_executor(None, self._parse_sync, file_path)
    
    def _parse_sync(self, file_path: str) -> ParsedDocument:
        try:
            import openpyxl
        except ImportError:
            raise ImportError("openpyxl not installed. Run: pip install openpyxl")
        
        wb = openpyxl.load_workbook(file_path, data_only=True)
        
        sheets = []
        all_text = []
        tables = []
        
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            
            sheet_text = []
            sheet_data = []
            
            for row in sheet.iter_rows(values_only=True):
                row_data = [str(cell) if cell is not None else '' for cell in row]
                
                if any(cell.strip() for cell in row_data):
                    sheet_text.append(' | '.join(row_data))
                    sheet_data.append(row_data)
            
            sheet_content = '\n'.join(sheet_text)
            all_text.append(f"## 工作表: {sheet_name}\n\n{sheet_content}")
            
            sheets.append({
                'sheet_name': sheet_name,
                'content': sheet_content,
                'row_count': sheet.max_row,
                'col_count': sheet.max_column,
            })
            
            if sheet_data:
                tables.append({
                    'sheet_name': sheet_name,
                    'data': sheet_data[:100],
                    'rows': len(sheet_data),
                    'cols': max(len(row) for row in sheet_data) if sheet_data else 0,
                })
        
        content = '\n\n'.join(all_text)
        
        metadata = {
            'filename': Path(file_path).name,
            'file_type': 'xlsx',
            'sheet_count': len(wb.sheetnames),
            'sheet_names': wb.sheetnames,
        }
        
        return ParsedDocument(
            content=content,
            metadata=metadata,
            tables=tables,
        )
    
    def supported_extensions(self) -> List[str]:
        return ['.xlsx', '.xls']
    
    def chunk_by_sheets(
        self,
        parsed_doc: ParsedDocument,
        chunk_size: int = 1000,
    ) -> List[ChunkResult]:
        chunks = []
        
        if parsed_doc.tables:
            for table in parsed_doc.tables:
                table_text = self._table_to_text(table)
                
                if len(table_text) <= chunk_size:
                    chunks.append(ChunkResult(
                        content=table_text,
                        token_count=self.count_tokens(table_text),
                        metadata={
                            'sheet_name': table['sheet_name'],
                            'is_table': True,
                        },
                    ))
                else:
                    text_chunks = self.chunk_text(table_text, chunk_size, 100)
                    for i, chunk in enumerate(text_chunks):
                        chunks.append(ChunkResult(
                            content=chunk,
                            token_count=self.count_tokens(chunk),
                            metadata={
                                'sheet_name': table['sheet_name'],
                                'is_table': True,
                                'chunk_index': i,
                            },
                        ))
        else:
            text_chunks = self.chunk_text(parsed_doc.content, chunk_size)
            for i, chunk in enumerate(text_chunks):
                chunks.append(ChunkResult(
                    content=chunk,
                    token_count=self.count_tokens(chunk),
                    metadata={'chunk_index': i},
                ))
        
        return chunks
    
    def _table_to_text(self, table: Dict) -> str:
        if not table.get('data'):
            return ''
        
        lines = [f"## 工作表: {table['sheet_name']}\n"]
        
        for row in table['data']:
            lines.append(' | '.join(row))
        
        return '\n'.join(lines)
